from __future__ import annotations

import csv
import json
import re
from copy import copy
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Iterable

import openpyxl
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import to_excel


ROOT = Path(__file__).resolve().parents[1]
HTML_TOOL_DIR = ROOT / "html导入工具"
MANUAL_STRUCTURE_PATH = HTML_TOOL_DIR / "v6_1_manual_2026_6_structure.json"
PRICE_TABLE_PATH = ROOT / "电商产品价格表-2026-06-24.xlsx"
OLD_CONFIG_PATH = HTML_TOOL_DIR / "report-config.js"
OUTPUT_WORKBOOK_PATH = HTML_TOOL_DIR / "2026年智能报量-v6.1.xlsx"
OUTPUT_CONFIG_PATH = HTML_TOOL_DIR / "report-config.js"
AUDIT_CSV_PATH = HTML_TOOL_DIR / "v6.1料号核对表.csv"
QA_JSON_PATH = HTML_TOOL_DIR / "v6_1_build_qa.json"


SOURCE_COLUMNS = {
    "storeName": "店铺名称",
    "paymentTime": "付款时间",
    "materialCode": "商品名称",
    "quantity": "订购数",
    "tradeStatus": "平台交易状态",
    "voidFlag": "作废",
    "refundFlag": "是否退款",
    "giftFlag": "赠品",
}


FILTERS = {
    "excludedTradeStatuses": ["取消交易", "交易取消", "交易关闭", "已取消"],
    "excludedVoidValues": ["是"],
    "excludedRefundValues": ["退款成功"],
    "excludedGiftValues": ["是"],
}


SHIFT = {
    "dayStart": "00:00",
    "dayEnd": "16:00",
    "nightStart": "16:00",
    "nightEnd": "24:00",
    "dayEndBelongsToDay": True,
}


TEMPLATE_CONFIG = {
    "firstDataColumn": 4,
    "dateRow": 1,
    "headerProductColumn": 3,
    "productTotalColumn": 1,
    "salesTotalColumn": 2,
    "productNameColumn": 3,
    "topQuantityCell": "B3",
    "topAmountCell": "C3",
    "topDailyRow": 3,
}
DATE_SHIFT_COLUMN_WIDTH = 9.0
SHIFT_DAY_FILL_COLOR = "FFEDEDED"
SHIFT_NIGHT_FILL_COLOR = "FFBDD7EE"


BLOCK_SOURCE_STORES_BY_DISPLAY = {
    "德达官方旗舰店": ["德达官方旗舰店", "京东1店", "德达官方旗舰店-国补"],
    "京东DEDAKJ旗舰店": ["德达旗舰店(京东)", "京东2店"],
    "DEDAKJ器械旗舰店": ["DEDAKJ器械旗舰店", "DEDAKJ器械旗舰店-国补", "京东3店"],
    "DEDAKJ个护健康旗舰店": ["DEDAKJ个护健康旗舰店", "京东8店"],
    "DEDAKJ自营旗舰店": ["DEDAKJ自营旗舰店", "京东6店"],
    "DEDAKJ保健器械旗舰店": ["DEDAKJ保健器械旗舰店", "京东5S店"],
    "天猫1店德达旗舰店": ["德达旗舰店-天猫", "天猫1店"],
    "天猫2店dedakj旗舰店": ["dedakj旗舰店", "天猫2店"],
    "天猫3店dedakj德达专卖店": ["dedakj德达专卖店"],
    "天猫6店德迩杰旗舰店": ["德迩杰旗舰店", "天猫6店"],
    "拼多多2店德达医疗旗舰店": ["德达医疗旗舰店", "拼多多2店"],
    "拼多多3店dedakj医疗旗舰店": ["DEDAKJ医疗器械官方旗舰店", "拼多多3店", "德达医疗器械旗舰店"],
    "抖音02店DEDAKJ个护健康旗舰店": ["德迩杰抖音店", "抖音-德迩杰官方旗舰店"],
    "抖音03店德达官方旗舰店": ["抖音德达制氧机"],
    "抖音05店dedakj医疗旗舰店": ["抖音DEDAKJ官方旗舰店", "抖音-德达医疗器械旗舰店"],
    "德达富氧": ["德达富氧", "deda9999"],
    "🔥💯👍（有赞平台：运营李宏生）": ["德达富氧", "deda9999"],
}


PRICE_TABLE_ALIAS_BY_ROW = {
    16: ["制氧机A1", "制氧机DH21-A1"],
    17: ["制氧机A1L"],
    18: ["制氧机Q1", "制氧机DD-Q1"],
    19: ["制氧机Q1W", "制氧机DD-Q1W"],
    20: ["制氧机Q2L"],
    21: ["制氧机1S01", "制氧机1S"],
    22: ["制氧机1SW01", "制氧机1SW"],
    23: ["制氧机2SW"],
    24: ["制氧机1SW PRO"],
    25: ["制氧机2SW PRO"],
    26: ["便携制氧机KH22-Y105 单电池", "便携制氧机Y105 单电池"],
    27: ["便携制氧机KH22-Y105 双电池", "便携制氧机Y105 双电池"],
    28: ["便携制氧机KH22-Y106 单电池", "便携制氧机Y106 单电池"],
    29: ["便携制氧机KH22-Y106 双电池", "便携制氧机Y106 双电池"],
    30: ["制氧机DE-01S无灭菌款", "制氧机DE-01S"],
    31: ["制氧机DE-Y300W"],
    32: ["制氧机Q3L", "制氧机DY22-Q3L"],
    33: ["制氧机Q5L", "制氧机DY22-Q5L"],
    34: ["制氧机Q5A", "制氧机DY22-Q5A"],
    35: ["制氧机Q5S", "制氧机DY22-Q5S"],
    36: ["制氧机Q5W", "制氧机DY22-Q5W"],
    37: ["制氧机Q10L", "制氧机DY22-Q10L"],
    38: ["便携制氧机MY-5C单电池", "便携制氧机MY-5C 单电池"],
    39: ["便携制氧机MY-5C双电池", "便携制氧机MY-5C 双电池"],
    40: ["便携制氧机YS-8V-7L单电池", "便携制氧机YS-8Y-7L单电池"],
    41: ["便携制氧机YS-8Y-7L双电池"],
    42: ["便携制氧机YS-8Y单电池", "便携制氧机YS-8Y 单电池"],
    43: ["便携制氧机YS-8Y双电池", "便携制氧机YS-8Y 双电池"],
    44: ["制氧机DE-Y3AW"],
    45: ["制氧机DE-Y5AW"],
    46: ["制氧机C1"],
    47: ["制氧机C1L"],
    48: ["制氧机Q1", "制氧机DD-Q1"],
    49: ["制氧机Q1W", "制氧机DD-Q1W"],
    50: ["制氧机Q2L"],
    51: ["制氧机DD-1A01"],
    52: ["制氧机DD-1LW01"],
    53: ["制氧机DD-2AW01"],
    54: ["便携制氧机DH21-V1"],
    55: ["便携制氧机DH21-V3"],
    56: ["便携制氧机DH21-V6"],
    57: ["制氧机DH21-A1"],
    58: ["制氧机DDH21-A1L", "制氧机DH21-A1L"],
    59: ["制氧机Y5A"],
    60: ["制氧机Y5L"],
    61: ["制氧机Y5S"],
    62: ["制氧机Y5W"],
    63: ["制氧机DY22-Q10L"],
    64: ["制氧机DY22-Q10W"],
    65: ["制氧机Y300W", "制氧机DE-Y300W"],
    66: ["制氧机Q3L", "制氧机DY22-Q3L"],
    67: ["制氧机Q5L", "制氧机DY22-Q5L"],
    68: ["制氧机Q5W", "制氧机DY22-Q5W"],
    69: ["制氧机Q10L", "制氧机DY22-Q10L"],
}


MONTH_DAYS = {
    1: 31,
    2: 28,
    3: 31,
    4: 30,
    5: 31,
    6: 30,
    7: 31,
    8: 31,
    9: 30,
    10: 31,
    11: 30,
    12: 31,
}


def log(action: str, module: str, sub_action: str) -> None:
    print(f"[build_v6_1_template.py][主线:{action}][{module}][{sub_action}]")


def normalize_text(value: object) -> str:
    return str(value or "").replace("\t", "").strip()


def normalize_key(value: str) -> str:
    text = strip_price(value)
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\s+", "", text)
    text = text.upper()
    text = text.replace("PRO", "PRO")
    return text


def strip_price(product_name: str) -> str:
    return re.sub(r"（[0-9]+(?:\.[0-9]+)?元）", "", normalize_text(product_name)).strip()


def split_codes(value: object) -> list[str]:
    text = normalize_text(value)
    if not text:
        return []
    parts = re.split(r"[，,、\s]+", text)
    return [part.strip() for part in parts if re.match(r"^[0-9]+\.", part.strip())]


def read_report_config(path: Path) -> dict:
    text = path.read_text(encoding="utf-8-sig")
    match = re.search(r"window\.REPORT_IMPORT_CONFIG\s*=\s*(\{.*\});\s*$", text, re.S)
    if not match:
        raise RuntimeError("无法解析 report-config.js")
    return json.loads(match.group(1))


def read_manual_structure() -> dict:
    return json.loads(MANUAL_STRUCTURE_PATH.read_text(encoding="utf-8-sig"))


def read_price_aliases() -> dict[str, list[str]]:
    workbook = openpyxl.load_workbook(PRICE_TABLE_PATH, data_only=True)
    worksheet = workbook["电商货盘"]
    alias_map: dict[str, list[str]] = {}
    for row_number, aliases in PRICE_TABLE_ALIAS_BY_ROW.items():
        codes = split_codes(worksheet.cell(row_number, 7).value)
        for alias in aliases:
            if codes:
                alias_map.setdefault(normalize_key(alias), [])
                alias_map[normalize_key(alias)] = unique(alias_map[normalize_key(alias)] + codes)
    return alias_map


def build_old_material_indexes(old_config: dict) -> tuple[dict[str, list[str]], dict[str, list[str]]]:
    by_name: dict[str, list[str]] = {}
    by_store_and_name: dict[str, list[str]] = {}
    for product_row in old_config.get("productRows", []):
        product_key = normalize_key(product_row.get("productName", ""))
        codes = product_row.get("materialCodes", [])
        if not product_key or not codes:
            continue
        by_name[product_key] = unique(by_name.get(product_key, []) + codes)
        for store in product_row.get("stores", []):
            key = f"{normalize_text(store)}||{product_key}"
            by_store_and_name[key] = unique(by_store_and_name.get(key, []) + codes)
    return by_name, by_store_and_name


@dataclass
class Block:
    index: int
    header_row: int
    summary_row: int
    store_row: int | None
    store_name: str
    operator_name: str
    products: list[dict]
    source_stores: list[str]
    included_in_top: bool


def parse_top_ranges(formula: str) -> set[int]:
    rows: set[int] = set()
    for start, end in re.findall(r"D(\d+):D(\d+)", formula or ""):
        rows.update(range(int(start), int(end) + 1))
    return rows


def build_blocks(manual: dict) -> list[Block]:
    markers = manual["markers"]
    products = manual["products"]
    top_rows = parse_top_ranges(manual["top"]["topAmountFormula"])
    blocks: list[Block] = []
    product_markers = [marker for marker in markers if marker["colE"] == "产品"]
    for index, marker in enumerate(product_markers, start=1):
        header_row = marker["row"]
        next_marker_rows = [item["row"] for item in markers if item["row"] > header_row and item["colE"] != "产品"]
        end_row = min(next_marker_rows) - 1 if next_marker_rows else manual["rows"]
        block_products = [product for product in products if header_row < product["row"] <= end_row]
        summary_row = header_row - 1
        operator_name = normalize_text(next((item["colE"] for item in markers if item["row"] == summary_row), ""))
        store_row = summary_row - 1
        store_name = normalize_text(next((item["colE"] for item in markers if item["row"] == store_row), ""))
        if store_name == "产品" or not store_name:
            store_row = None
            store_name = operator_name
        source_stores = BLOCK_SOURCE_STORES_BY_DISPLAY.get(store_name.strip(), [store_name.strip()])
        blocks.append(
            Block(
                index=index,
                header_row=header_row,
                summary_row=summary_row,
                store_row=store_row,
                store_name=store_name,
                operator_name=operator_name,
                products=block_products,
                source_stores=source_stores,
                included_in_top=any(product["row"] in top_rows for product in block_products),
            )
        )
    return blocks


def unique(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    output: list[str] = []
    for value in values:
        text = normalize_text(value)
        if text and text not in seen:
            seen.add(text)
            output.append(text)
    return output


def resolve_material_codes(
    product_name: str,
    source_stores: list[str],
    old_by_name: dict[str, list[str]],
    old_by_store_and_name: dict[str, list[str]],
    price_aliases: dict[str, list[str]],
) -> tuple[list[str], str, list[str], str]:
    product_key = normalize_key(product_name)
    old_codes: list[str] = []
    for store in source_stores:
        old_codes += old_by_store_and_name.get(f"{normalize_text(store)}||{product_key}", [])
    old_codes = unique(old_codes)
    global_old_codes = old_by_name.get(product_key, [])
    price_codes = price_aliases.get(product_key, [])

    if old_codes:
        selected_codes = unique(old_codes + price_codes)
        source = "v6.0同店同型号"
        if price_codes and set(price_codes) - set(old_codes):
            source = "v6.0同店同型号+价格表补齐"
        return selected_codes, source, price_codes, compare_code_source(selected_codes, price_codes)
    if global_old_codes:
        selected_codes = unique(global_old_codes + price_codes)
        source = "v6.0同型号跨店继承"
        if price_codes and set(price_codes) - set(global_old_codes):
            source = "v6.0同型号跨店继承+价格表补齐"
        return selected_codes, source, price_codes, compare_code_source(selected_codes, price_codes)
    if price_codes:
        return price_codes, "最新价格表", price_codes, "价格表补齐"
    return [], "未找到料号", [], "需人工核对"


def compare_code_source(selected_codes: list[str], price_codes: list[str]) -> str:
    # 多料号继承时，只要已经覆盖价格表料号，就保留多料号，避免把历史老料号误删。
    if not price_codes:
        return "价格表无同名料号"
    selected_code_set = set(selected_codes)
    price_code_set = set(price_codes)
    if selected_code_set == price_code_set:
        return "一致"
    if price_code_set.issubset(selected_code_set):
        return "已覆盖价格表料号"
    return f"待核对：已选{';'.join(selected_codes)} / 价格表{';'.join(price_codes)}"


def col_name(column_number: int) -> str:
    return get_column_letter(column_number)


def last_data_col(day_count: int) -> int:
    return 3 + day_count * 2


def day_col(day_index: int) -> int:
    return 4 + (day_index - 1) * 2


def night_col(day_index: int) -> int:
    return day_col(day_index) + 1


def apply_date_column_widths(worksheet, day_count: int) -> None:
    # 日期区按班次统一列宽，避免白班夜班长期维护成两套规则。
    for day in range(1, day_count + 1):
        worksheet.column_dimensions[col_name(day_col(day))].width = DATE_SHIFT_COLUMN_WIDTH
        worksheet.column_dimensions[col_name(night_col(day))].width = DATE_SHIFT_COLUMN_WIDTH


def product_amount_formula(row_number: int) -> str:
    return (
        f'=+A{row_number}*MID(C{row_number},FIND("（",C{row_number})+1,'
        f'FIND("）",C{row_number})-FIND("（",C{row_number})-2)'
    )


def make_formula_sum_cells(column: int, rows: list[int]) -> str:
    cells = [f"{col_name(column)}{row}" for row in rows]
    return f"=SUM({','.join(cells)})" if cells else "=0"


def make_formula_sum_ranges(column: int, row_ranges: list[tuple[int, int]]) -> str:
    # 顶部月销售额按店铺产品区间汇总，避免超长单格列表被 Excel 清空。
    ranges = [f"{col_name(column)}{start}:{col_name(column)}{end}" for start, end in row_ranges if start <= end]
    return f"=SUM({','.join(ranges)})" if ranges else "=0"


def merge_day_header_pairs(worksheet, row_number: int, day_count: int) -> None:
    # 日期标题类行按 v6.0 结构合并白班/夜班两列，保留明细写入列不合并。
    for day in range(1, day_count + 1):
        worksheet.merged_cells.add(
            f"{col_name(day_col(day))}{row_number}:{col_name(night_col(day))}{row_number}"
        )


def apply_base_style(workbook: Workbook) -> dict[str, object]:
    font = Font(name="微软雅黑", size=10, color="1F2933")
    bold_font = Font(name="微软雅黑", size=10, bold=True, color="111827")
    white_bold_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    small_note_font = Font(name="微软雅黑", size=9, color="6B7280")
    border = Border(
        left=Side(style="thin", color="D9E2EC"),
        right=Side(style="thin", color="D9E2EC"),
        top=Side(style="thin", color="D9E2EC"),
        bottom=Side(style="thin", color="D9E2EC"),
    )
    return {
        "font": font,
        "bold_font": bold_font,
        "white_bold_font": white_bold_font,
        "small_note_font": small_note_font,
        "border": border,
        "top_fill": PatternFill("solid", fgColor="D9EAF7"),
        "store_fill": PatternFill("solid", fgColor="1F4E79"),
        "operator_fill": PatternFill("solid", fgColor="E8F1F8"),
        "header_fill": PatternFill("solid", fgColor="DDEBF7"),
        "product_fill": PatternFill("solid", fgColor="FFFFFF"),
        "total_fill": PatternFill("solid", fgColor="E2F0D9"),
        "zero_fill": PatternFill("solid", fgColor="FFFFFF"),
        "day_shift_fill": PatternFill("solid", fgColor=SHIFT_DAY_FILL_COLOR),
        "night_shift_fill": PatternFill("solid", fgColor=SHIFT_NIGHT_FILL_COLOR),
        "center": Alignment(horizontal="center", vertical="center"),
        "left": Alignment(horizontal="left", vertical="center"),
        "right": Alignment(horizontal="right", vertical="center"),
    }


def style_row(worksheet, row_number: int, max_col: int, style: dict[str, object], role: str) -> None:
    for column in range(1, max_col + 1):
        cell = worksheet.cell(row_number, column)
        cell.border = copy(style["border"])
        cell.alignment = copy(style["center"])
        cell.font = copy(style["font"])
        if role == "top":
            cell.fill = copy(style["top_fill"])
            cell.font = copy(style["bold_font"])
        elif role == "store":
            cell.fill = copy(style["store_fill"])
            cell.font = copy(style["white_bold_font"])
        elif role == "operator":
            cell.fill = copy(style["operator_fill"])
            cell.font = copy(style["bold_font"])
        elif role == "header":
            cell.fill = copy(style["header_fill"])
            cell.font = copy(style["bold_font"])
        elif role == "product":
            cell.fill = copy(style["product_fill"])
            if column == 3:
                cell.alignment = copy(style["left"])
            if column in (1, 2):
                cell.alignment = copy(style["right"])
        elif role == "blank":
            cell.fill = copy(style["zero_fill"])


def apply_shift_column_fills(worksheet, day_count: int, max_row: int, style: dict[str, object]) -> None:
    # 这个函数用于把白班/夜班颜色集中校准，避免复用模板时出现有的标色、有的没标色。
    for row_number in range(4, max_row + 1):
        for day in range(1, day_count + 1):
            day_cell = worksheet.cell(row_number, day_col(day))
            night_cell = worksheet.cell(row_number, night_col(day))
            if not isinstance(day_cell, MergedCell):
                day_cell.fill = copy(style["day_shift_fill"])
            if not isinstance(night_cell, MergedCell):
                night_cell.fill = copy(style["night_shift_fill"])


def write_month_sheet(
    workbook: Workbook,
    sheet_name: str,
    month: int,
    blocks: list[Block],
    manual_rows: int,
    style: dict[str, object],
) -> None:
    worksheet = workbook.create_sheet(sheet_name)
    days = MONTH_DAYS[month]
    max_col = last_data_col(days)
    top_summary_rows = [block.summary_row for block in blocks if block.included_in_top]
    top_product_ranges = [
        (block.products[0]["row"], block.products[-1]["row"])
        for block in blocks
        if block.included_in_top and block.products
    ]

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "D7"
    worksheet.row_dimensions[1].hidden = True
    worksheet.row_dimensions[1].height = 0
    worksheet.column_dimensions["A"].width = 10
    worksheet.column_dimensions["B"].width = 15
    worksheet.column_dimensions["C"].width = 38
    apply_date_column_widths(worksheet, days)

    worksheet["A1"] = f"程序日期定位行，勿删；用于导入工具识别{month}月白班/夜班列。"
    worksheet["A1"].font = copy(style["small_note_font"])
    for day in range(1, days + 1):
        serial = to_excel(date(2026, month, day))
        worksheet.cell(1, day_col(day), serial)
        worksheet.cell(1, night_col(day), serial)
        worksheet.cell(2, day_col(day), serial)
        worksheet.cell(2, night_col(day), None)
        worksheet.cell(2, day_col(day)).number_format = "m/d"

    style_row(worksheet, 2, max_col, style, "top")
    worksheet["A2"] = "产品"
    worksheet["B2"] = "当月销量"
    worksheet["C2"] = "【国内电商】月销售额"
    style_row(worksheet, 3, max_col, style, "top")
    worksheet["A3"] = "制氧机"
    worksheet["B3"] = f"=SUM(D3:{col_name(max_col)}3)"
    worksheet["C3"] = make_formula_sum_ranges(2, top_product_ranges)
    for day in range(1, days + 1):
        worksheet.cell(3, day_col(day), make_formula_sum_cells(night_col(day), top_summary_rows))
        worksheet.cell(3, day_col(day)).number_format = "General"
        worksheet.cell(3, night_col(day), None)

    merge_day_header_pairs(worksheet, 2, days)
    merge_day_header_pairs(worksheet, 3, days)

    for row_number in range(4, manual_rows + 1):
        style_row(worksheet, row_number, max_col, style, "blank")

    for block in blocks:
        if block.store_row:
            style_row(worksheet, block.store_row, max_col, style, "store")
            worksheet.cell(block.store_row, 3, block.store_name)
            for day in range(1, days + 1):
                worksheet.cell(block.store_row, day_col(day), to_excel(date(2026, month, day)))
                worksheet.cell(block.store_row, day_col(day)).number_format = "m/d"
            merge_day_header_pairs(worksheet, block.store_row, days)
        style_row(worksheet, block.summary_row, max_col, style, "operator")
        worksheet.cell(block.summary_row, 3, block.operator_name)
        for day in range(1, days + 1):
            first_row = block.products[0]["row"] if block.products else block.header_row + 1
            last_row = block.products[-1]["row"] if block.products else block.header_row
            worksheet.cell(block.summary_row, day_col(day), "本店汇总")
            worksheet.cell(
                block.summary_row,
                night_col(day),
                f"=SUM({col_name(day_col(day))}{first_row}:{col_name(night_col(day))}{last_row})",
            )
            worksheet.cell(block.summary_row, night_col(day)).number_format = "General"

        style_row(worksheet, block.header_row, max_col, style, "header")
        worksheet.cell(block.header_row, 1, "汇总")
        worksheet.cell(block.header_row, 2, "销售额累计")
        worksheet.cell(block.header_row, 3, "产品")
        for day in range(1, days + 1):
            worksheet.cell(block.header_row, day_col(day), "白班")
            worksheet.cell(block.header_row, night_col(day), "夜班")

        for product in block.products:
            row_number = product["row"]
            style_row(worksheet, row_number, max_col, style, "product")
            worksheet.cell(row_number, 1, f"=SUM(D{row_number}:{col_name(max_col)}{row_number})")
            worksheet.cell(row_number, 2, product_amount_formula(row_number))
            worksheet.cell(row_number, 3, product["productName"])
            for day in range(1, days + 1):
                worksheet.cell(row_number, day_col(day), 0)
                worksheet.cell(row_number, night_col(day), 0)

    apply_shift_column_fills(worksheet, days, manual_rows, style)

    for row_number in range(1, manual_rows + 1):
        worksheet.row_dimensions[row_number].height = 20
    worksheet.row_dimensions[2].height = 24
    worksheet.row_dimensions[3].height = 24


def build_config_and_audit(
    blocks: list[Block],
    old_config: dict,
    price_aliases: dict[str, list[str]],
) -> tuple[dict, list[dict]]:
    old_by_name, old_by_store_and_name = build_old_material_indexes(old_config)
    product_rows: list[dict] = []
    audit_rows: list[dict] = []
    for block in blocks:
        for product in block.products:
            codes, source, price_codes, note = resolve_material_codes(
                product["productName"],
                block.source_stores,
                old_by_name,
                old_by_store_and_name,
                price_aliases,
            )
            product_rows.append(
                {
                    "row": product["row"],
                    "productName": product["productName"],
                    "stores": block.source_stores,
                    "materialCodes": codes,
                }
            )
            audit_rows.append(
                {
                    "店铺序号": block.index,
                    "展示店铺": block.store_name,
                    "订单店铺": "；".join(block.source_stores),
                    "行号": product["row"],
                    "型号价格": product["productName"],
                    "手工价格": product["price"],
                    "料号来源": source,
                    "已填料号": "；".join(codes),
                    "价格表同名料号": "；".join(price_codes),
                    "核对状态": note,
                }
            )

    material_preference: dict[str, str] = {}
    for product_row in product_rows:
        preferred_name = normalize_key(product_row["productName"])
        for material_code in product_row["materialCodes"]:
            material_preference.setdefault(material_code, preferred_name)

    config = {
        "version": "6.1.0",
        "generatedFrom": "2026年【报量】表.xlsx / 2026-6 + 电商产品价格表-2026-06-24.xlsx",
        "supportedSheetMonths": list(range(1, 13)),
        "sourceColumns": SOURCE_COLUMNS,
        "filters": FILTERS,
        "shift": SHIFT,
        "template": TEMPLATE_CONFIG,
        "productRows": product_rows,
        "materialCodePreferredProductName": material_preference,
    }
    return config, audit_rows


def write_report_config(config: dict) -> None:
    text = (
        "// 该文件由 v6.1 构建脚本生成，目的是真实记录手工6月结构和料号映射，避免人工抄错。\n"
        "window.REPORT_IMPORT_CONFIG = "
        + json.dumps(config, ensure_ascii=False, indent=2)
        + ";\n"
    )
    OUTPUT_CONFIG_PATH.write_text(text, encoding="utf-8")


def write_audit_csv(rows: list[dict]) -> None:
    fieldnames = ["店铺序号", "展示店铺", "订单店铺", "行号", "型号价格", "手工价格", "料号来源", "已填料号", "价格表同名料号", "核对状态"]
    with AUDIT_CSV_PATH.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def scan_formula_errors(workbook: Workbook) -> list[dict]:
    errors: list[dict] = []
    pattern = re.compile(r"#REF!|#DIV/0!|#VALUE!|#NAME\?|#N/A")
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and pattern.search(cell.value):
                    errors.append({"sheet": worksheet.title, "cell": cell.coordinate, "value": cell.value})
    return errors


def main() -> None:
    log("读取", "输入", "加载手工结构、旧配置、价格表")
    manual = read_manual_structure()
    old_config = read_report_config(OLD_CONFIG_PATH)
    price_aliases = read_price_aliases()
    blocks = build_blocks(manual)
    if len(blocks) != 16:
        raise RuntimeError(f"手工表店铺块数量异常：{len(blocks)}，预期16")

    log("生成", "配置", "建立料号映射和核对表")
    config, audit_rows = build_config_and_audit(blocks, old_config, price_aliases)

    log("生成", "工作簿", "创建v6.1全年模板")
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    style = apply_base_style(workbook)
    for month in range(1, 13):
        write_month_sheet(workbook, f"2026-{month}", month, blocks, manual["rows"], style)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    formula_errors = scan_formula_errors(workbook)
    if formula_errors:
        raise RuntimeError(f"生成模板包含公式错误：{formula_errors[:5]}")

    workbook.save(OUTPUT_WORKBOOK_PATH)
    write_report_config(config)
    write_audit_csv(audit_rows)

    qa = {
        "workbook": str(OUTPUT_WORKBOOK_PATH),
        "config": str(OUTPUT_CONFIG_PATH),
        "auditCsv": str(AUDIT_CSV_PATH),
        "sheetCount": 12,
        "blockCount": len(blocks),
        "manualProductRows": sum(len(block.products) for block in blocks),
        "configProductRows": len(config["productRows"]),
        "rowsWithoutMaterialCodes": [
            {"row": row["行号"], "product": row["型号价格"], "store": row["展示店铺"]}
            for row in audit_rows
            if not row["已填料号"]
        ],
        "formulaErrorCount": len(formula_errors),
    }
    QA_JSON_PATH.write_text(json.dumps(qa, ensure_ascii=False, indent=2), encoding="utf-8")
    log("完成", "输出", f"模板={OUTPUT_WORKBOOK_PATH.name} 配置行={len(config['productRows'])} 无料号={len(qa['rowsWithoutMaterialCodes'])}")


if __name__ == "__main__":
    main()
