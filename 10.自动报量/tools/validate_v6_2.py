from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import to_excel


ROOT = Path(__file__).resolve().parents[1]
HTML_TOOL_DIR = ROOT / "html导入工具"
WORKBOOK_PATH = HTML_TOOL_DIR / "2026年智能报量-v6.2.xlsx"
# 本文件与 tools/validate_v6_1.py 为近逐字镜像（仅目标工作簿/QA路径与班次颜色校验差异），
# 修改校验逻辑时务必同步另一份（同#543先例，跨版本工具不合并只注释）。
MANUAL_STRUCTURE_PATH = HTML_TOOL_DIR / "v6_1_manual_2026_6_structure.json"
CONFIG_PATH = HTML_TOOL_DIR / "report-config.js"
QA_PATH = HTML_TOOL_DIR / "v6_2_validation_qa.json"
DATE_SHIFT_COLUMN_WIDTH = 9.0
DATE_COLUMN_WIDTH_TOLERANCE = 0.01

MONTH_DAYS = {
    1: 31,
    2: 28,
    3: 31,
    4: 30,
    5: 31,
    6: 30,
    7: 31,
    8: 31,
    9: 30,
    10: 31,
    11: 30,
    12: 31,
}


def log(action: str, module: str, sub_action: str) -> None:
    print(f"[validate_v6_2.py][主线:{action}][{module}][{sub_action}]")


def normalize_excel_serial(value):
    # 样式复用后Excel会把隐藏日期定位行按日期读出，这里统一转回序列号再比较。
    if isinstance(value, datetime):
        return to_excel(value)
    if isinstance(value, date):
        return to_excel(value)
    return value


def is_same_excel_serial(value, expected_serial: float) -> bool:
    # 日期定位只关心底层序列号一致，避免被显示格式影响验收结果。
    normalized_value = normalize_excel_serial(value)
    try:
        return abs(float(normalized_value) - float(expected_serial)) < 0.000001
    except (TypeError, ValueError):
        return normalized_value == expected_serial


def read_config() -> dict:
    text = CONFIG_PATH.read_text(encoding="utf-8-sig")
    match = re.search(r"window\.REPORT_IMPORT_CONFIG\s*=\s*(\{.*\});\s*$", text, re.S)
    if not match:
        raise RuntimeError("report-config.js 格式异常")
    return json.loads(match.group(1))


def last_data_col(day_count: int) -> int:
    return 3 + day_count * 2


def day_col(day_index: int) -> int:
    return 4 + (day_index - 1) * 2


def night_col(day_index: int) -> int:
    return day_col(day_index) + 1


def col_name(column_number: int) -> str:
    return get_column_letter(column_number)


def make_formula_sum_cells(column: int, rows: list[int]) -> str:
    # 按指定单元格逐个汇总，用于顶部每日汇总行。
    cells = [f"{col_name(column)}{row}" for row in rows]
    return f"=SUM({','.join(cells)})" if cells else "=0"


def make_formula_sum_ranges(column: int, row_ranges: list[tuple[int, int]]) -> str:
    # 按店铺产品区间汇总，避免顶部月销售额公式过长或漏块。
    ranges = [f"{col_name(column)}{start}:{col_name(column)}{end}" for start, end in row_ranges if start <= end]
    return f"=SUM({','.join(ranges)})" if ranges else "=0"


def parse_top_ranges(formula: str) -> set[int]:
    rows: set[int] = set()
    for start, end in re.findall(r"D(\d+):D(\d+)", formula or ""):
        rows.update(range(int(start), int(end) + 1))
    return rows


def expected_blocks(manual: dict) -> list[dict]:
    # 从手工结构里还原店铺块，验证 v6.2 没有漏掉店铺标题和汇总行。
    markers = manual["markers"]
    products = manual["products"]
    top_rows = parse_top_ranges(manual["top"]["topAmountFormula"])
    product_markers = [marker for marker in markers if marker["colE"] == "产品"]
    blocks: list[dict] = []
    for marker in product_markers:
        header_row = marker["row"]
        next_marker_rows = [item["row"] for item in markers if item["row"] > header_row and item["colE"] != "产品"]
        end_row = min(next_marker_rows) - 1 if next_marker_rows else manual["rows"]
        block_products = [product for product in products if header_row < product["row"] <= end_row]
        summary_row = header_row - 1
        store_row = summary_row - 1
        store_name = strip_text(next((item["colE"] for item in markers if item["row"] == store_row), ""))
        if store_name == "产品" or not store_name:
            store_row = None
        blocks.append(
            {
                "headerRow": header_row,
                "summaryRow": summary_row,
                "storeRow": store_row,
                "storeName": store_name,
                "productRows": [product["row"] for product in block_products],
                "includedInTop": any(product["row"] in top_rows for product in block_products),
            }
        )
    return blocks


def strip_text(value: str) -> str:
    return str(value or "").replace("\u3000", " ").strip()


def is_merged_pair(worksheet, row_number: int, start_column: int, end_column: int) -> bool:
    # 日期标题合并必须精确覆盖白班/夜班两列，不能多合并或少合并。
    return any(
        merged_range.min_row == row_number
        and merged_range.max_row == row_number
        and merged_range.min_col == start_column
        and merged_range.max_col == end_column
        for merged_range in worksheet.merged_cells.ranges
    )


def is_same_width(actual_width: float | None, expected_width: float) -> bool:
    # Excel 和 openpyxl 对列宽存在微小浮点差异，验收允许 0.01 以内误差。
    return actual_width is not None and abs(float(actual_width) - expected_width) <= DATE_COLUMN_WIDTH_TOLERANCE


def price_of(product_name: str) -> str:
    match = re.search(r"（([0-9]+(?:\.[0-9]+)?)元）", str(product_name or ""))
    return match.group(1) if match else ""


def validate() -> dict:
    log("读取", "输入", "加载手工结构、v6.2模板和配置")
    manual = json.loads(MANUAL_STRUCTURE_PATH.read_text(encoding="utf-8-sig"))
    config = read_config()
    workbook = openpyxl.load_workbook(WORKBOOK_PATH, data_only=False)
    errors: list[str] = []
    warnings: list[str] = []

    expected_sheets = [f"2026-{month}" for month in range(1, 13)]
    if workbook.sheetnames != expected_sheets:
        errors.append(f"工作表顺序异常：{workbook.sheetnames}")

    manual_products = {item["row"]: item for item in manual["products"]}
    config_rows = {item["row"]: item for item in config["productRows"]}
    blocks = expected_blocks(manual)
    top_summary_rows = [block["summaryRow"] for block in blocks if block["includedInTop"]]
    top_product_ranges = [
        (block["productRows"][0], block["productRows"][-1])
        for block in blocks
        if block["includedInTop"] and block["productRows"]
    ]
    if len(config_rows) != len(manual_products):
        errors.append(f"配置产品行数量异常：config={len(config_rows)} manual={len(manual_products)}")
    if "deda9999" not in config_rows.get(385, {}).get("stores", []):
        errors.append("德达富氧店铺别名缺失：R385 stores 未包含 deda9999")

    for row_number, manual_product in manual_products.items():
        config_product = config_rows.get(row_number)
        if not config_product:
            errors.append(f"配置缺少产品行：{row_number} {manual_product['productName']}")
            continue
        if config_product["productName"] != manual_product["productName"]:
            errors.append(f"配置产品名不一致：R{row_number} {config_product['productName']} != {manual_product['productName']}")
        if not config_product.get("materialCodes"):
            warnings.append(f"配置产品行无料号：R{row_number} {manual_product['productName']}")

    formula_error_pattern = re.compile(r"#REF!|#DIV/0!|#VALUE!|#NAME\?|#N/A")
    external_ref_pattern = re.compile(r"\[[^\]]+\]")

    for month in range(1, 13):
        sheet_name = f"2026-{month}"
        worksheet = workbook[sheet_name]
        days = MONTH_DAYS[month]
        max_col = last_data_col(days)
        if worksheet.row_dimensions[1].hidden is not True:
            warnings.append(f"{sheet_name} 第1行未隐藏")
        for day in range(1, days + 1):
            expected_serial = to_excel(date(2026, month, day))
            day_value = worksheet.cell(1, day_col(day)).value
            night_value = worksheet.cell(1, night_col(day)).value
            if not is_same_excel_serial(day_value, expected_serial) or not is_same_excel_serial(night_value, expected_serial):
                errors.append(f"{sheet_name} 日期定位异常：第{day}天 {day_value}/{night_value} != {expected_serial}")
            day_width = worksheet.column_dimensions[col_name(day_col(day))].width
            night_width = worksheet.column_dimensions[col_name(night_col(day))].width
            if not is_same_width(day_width, DATE_SHIFT_COLUMN_WIDTH):
                errors.append(f"{sheet_name} 白班列宽异常：{col_name(day_col(day))}={day_width}")
            if not is_same_width(night_width, DATE_SHIFT_COLUMN_WIDTH):
                errors.append(f"{sheet_name} 夜班列宽异常：{col_name(night_col(day))}={night_width}")
            if not is_merged_pair(worksheet, 2, day_col(day), night_col(day)):
                errors.append(f"{sheet_name} 日期标题行未合并：R2 {col_name(day_col(day))}:{col_name(night_col(day))}")
            if not is_merged_pair(worksheet, 3, day_col(day), night_col(day)):
                errors.append(f"{sheet_name} 顶部日汇总行未合并：R3 {col_name(day_col(day))}:{col_name(night_col(day))}")
            expected_daily_formula = make_formula_sum_cells(night_col(day), top_summary_rows)
            if worksheet.cell(3, day_col(day)).value != expected_daily_formula:
                errors.append(f"{sheet_name} 顶部日汇总公式异常：{col_name(day_col(day))}3={worksheet.cell(3, day_col(day)).value}")
            if worksheet.cell(3, day_col(day)).number_format != "General":
                errors.append(f"{sheet_name} 顶部日汇总格式异常：{col_name(day_col(day))}3={worksheet.cell(3, day_col(day)).number_format}")
        for column in range(max_col + 1, max_col + 5):
            if worksheet.cell(1, column).value not in (None, ""):
                errors.append(f"{sheet_name} 日期定位多余值：{col_name(column)}1={worksheet.cell(1, column).value}")

        expected_top_amount_formula = make_formula_sum_ranges(2, top_product_ranges)
        if worksheet["C3"].value != expected_top_amount_formula:
            errors.append(f"{sheet_name} C3月销售额公式异常：{worksheet['C3'].value} != {expected_top_amount_formula}")

        for block in blocks:
            store_row = block["storeRow"]
            product_rows = block["productRows"]
            if store_row:
                if worksheet.cell(store_row, 3).value != block["storeName"]:
                    errors.append(f"{sheet_name} R{store_row} 店铺标题异常：{worksheet.cell(store_row, 3).value} != {block['storeName']}")
                for day in range(1, days + 1):
                    if not is_merged_pair(worksheet, store_row, day_col(day), night_col(day)):
                        errors.append(f"{sheet_name} R{store_row} 店铺日期标题未合并：{col_name(day_col(day))}:{col_name(night_col(day))}")
            if product_rows:
                first_row = product_rows[0]
                last_row = product_rows[-1]
                for day in range(1, days + 1):
                    expected_summary_formula = f"=SUM({col_name(day_col(day))}{first_row}:{col_name(night_col(day))}{last_row})"
                    if worksheet.cell(block["summaryRow"], day_col(day)).value != "本店汇总":
                        errors.append(
                            f"{sheet_name} R{block['summaryRow']} 本店汇总标题异常："
                            f"{col_name(day_col(day))}={worksheet.cell(block['summaryRow'], day_col(day)).value}"
                        )
                    if worksheet.cell(block["summaryRow"], night_col(day)).value != expected_summary_formula:
                        errors.append(
                            f"{sheet_name} R{block['summaryRow']} 本店汇总公式异常："
                            f"{col_name(night_col(day))}={worksheet.cell(block['summaryRow'], night_col(day)).value}"
                        )
                    if worksheet.cell(block["summaryRow"], night_col(day)).number_format != "General":
                        errors.append(
                            f"{sheet_name} R{block['summaryRow']} 本店汇总格式异常："
                            f"{col_name(night_col(day))}={worksheet.cell(block['summaryRow'], night_col(day)).number_format}"
                        )

        for row_number, manual_product in manual_products.items():
            product_name = worksheet.cell(row_number, 3).value
            if product_name != manual_product["productName"]:
                errors.append(f"{sheet_name} R{row_number} 产品不一致：{product_name} != {manual_product['productName']}")
            if price_of(str(product_name)) != manual_product["price"]:
                errors.append(f"{sheet_name} R{row_number} 价格不一致：{product_name} / 手工{manual_product['price']}")
            expected_qty_formula = f"=SUM(D{row_number}:{col_name(max_col)}{row_number})"
            if worksheet.cell(row_number, 1).value != expected_qty_formula:
                errors.append(f"{sheet_name} R{row_number} 销量公式异常：{worksheet.cell(row_number, 1).value}")
            if "MID(C" not in str(worksheet.cell(row_number, 2).value):
                errors.append(f"{sheet_name} R{row_number} 销售额公式异常：{worksheet.cell(row_number, 2).value}")
            for column in range(4, max_col + 1):
                if worksheet.cell(row_number, column).value not in (0, None):
                    errors.append(f"{sheet_name} R{row_number} 模板销量未清零：{col_name(column)}={worksheet.cell(row_number, column).value}")
                    break

        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str):
                    if formula_error_pattern.search(value):
                        errors.append(f"{sheet_name}!{cell.coordinate} 公式错误：{value}")
                    if external_ref_pattern.search(value):
                        errors.append(f"{sheet_name}!{cell.coordinate} 外部引用：{value}")

    qa = {
        "workbook": str(WORKBOOK_PATH),
        "config": str(CONFIG_PATH),
        "sheetCount": len(workbook.sheetnames),
        "configVersion": config.get("version"),
        "supportedSheetMonths": config.get("supportedSheetMonths"),
        "manualProductRows": len(manual_products),
        "configProductRows": len(config_rows),
        "errors": errors[:200],
        "errorCount": len(errors),
        "warnings": warnings[:200],
        "warningCount": len(warnings),
    }
    QA_PATH.write_text(json.dumps(qa, ensure_ascii=False, indent=2), encoding="utf-8")
    return qa


def main() -> None:
    qa = validate()
    log("完成", "验收", f"错误={qa['errorCount']} 警告={qa['warningCount']} 输出={QA_PATH.name}")
    if qa["errorCount"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
