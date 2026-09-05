"""样表列顺序、外部文本写盘安全及中途停止保存。"""
from pathlib import Path
import tempfile
import threading
import itertools
import json
import unittest
from unittest import mock

from openpyxl import load_workbook
import jd_shops as jd
from jd_session import JdPageReader, VerificationTimeout, CollectionStopped
from jd_fields import labeled_fields, COMPANY_FIELDS, SCORE_FIELDS


class JdShopTests(unittest.TestCase):
    def test_ids_must_match_and_remain_distinct(self):
        html = '<title>测试店 - 京东</title><input id="shop_id" value="123"><input id="vender_id" value="456">'
        row = jd.parse_shop_page(html, 'https://mall.jd.com/index-123.html')
        self.assertEqual((row['店铺ID'], row['VenderId']), ('123', '456'))
        with self.assertRaises(ValueError):
            jd.parse_shop_page(html, 'https://mall.jd.com/index-456.html')
        for html in ('<title>京东验证</title>', '<title>京东-欢迎登录</title>'):
            with self.assertRaises(ValueError):
                jd.parse_shop_page(html, 'https://mall.jd.com/index-123.html')

    def test_template_headers_and_text_roundtrip(self):
        reference = Path(__file__).with_name('京东耳机商家信息_20260720154222.xlsx')
        book = load_workbook(reference, read_only=True)
        try:
            expected = list(next(book.worksheets[0].iter_rows(values_only=True)))
            while expected and expected[-1] is None:
                expected.pop()
            expected = tuple(expected)
        finally:
            book.close()
        row = dict.fromkeys(jd.FIELDS, '')
        row.update({'店铺名': '=1+1', '店铺链接': 'https://mall.jd.com/index-123.html',
                    '店铺ID': '123', 'VenderId': '456', '电话': '00123456789'})
        with tempfile.TemporaryDirectory() as directory:
            path, report = jd.export_rows([row], [{'原因': '测试'}], directory)
            book = load_workbook(path, data_only=False)
            try:
                self.assertEqual(tuple(c.value for c in book.active[1]), expected)
                self.assertEqual(book.active['A2'].data_type, 's')
                self.assertEqual(book.active['A2'].value, '=1+1')
                self.assertEqual(book.active['I2'].value, '00123456789')
            finally:
                book.close()
            self.assertTrue(report.exists())
            self.assertEqual(jd.read_shop_urls(path), ['https://mall.jd.com/index-123.html'])

    def test_stop_saves_completed_shop(self):
        stop = threading.Event()
        url = 'https://mall.jd.com/index-123.html'
        response = {'url': url, 'html': '<title>Test - JD</title>', 'text': ''}
        row = dict.fromkeys(jd.FIELDS, '')
        row.update({'店铺名': '测试店', '店铺链接': url, '店铺ID': '123', 'VenderId': '456'})
        def parse(*args):
            stop.set()
            return row
        with tempfile.TemporaryDirectory() as directory, \
             mock.patch.object(jd, 'read_shop_urls', return_value=[url, 'https://mall.jd.com/index-789.html']), \
             mock.patch.object(jd, 'JdPageReader') as session, \
             mock.patch.object(jd, 'enrich_shop'), \
             mock.patch.object(jd, 'parse_shop_page', side_effect=parse):
            session.return_value.__enter__.return_value.read.return_value = response
            saved = jd.run_shops(outdir=directory, stop_event=stop)
            self.assertEqual(len(saved), 1)
            self.assertEqual(len(list(Path(directory).glob('*.xlsx'))), 1)
            self.assertEqual(session.return_value.__enter__.return_value.read.call_count, 1)

    def test_labels_do_not_turn_recruiter_or_footer_into_business_fields(self):
        values, conflicts = labeled_fields('招聘者：张某\n工作地点：杭州\n售后服务\n退换货政策\n公司名称：甲公司\n法人：张三\n电话：010-12345678', COMPANY_FIELDS + SCORE_FIELDS)
        self.assertEqual(values, {'公司名': '甲公司', '法人': '张三', '电话': '010-12345678'})
        self.assertFalse(conflicts)
        values, conflicts = labeled_fields('公司名称：甲公司\n公司名称：乙公司', COMPANY_FIELDS)
        self.assertNotIn('公司名', values)
        self.assertEqual(conflicts, ['公司名'])

    def test_enrichment_records_sources(self):
        row = dict.fromkeys(jd.FIELDS, '')
        row.update({'店铺链接': 'https://mall.jd.com/index-123.html', '店铺ID': '123'})
        page = {'text': '', 'html': '<input id="pageInstance_appId" value="99">'}
        header = '<li><span>商品评价</span><b>7.2</b></li><li>物流履约：8.1</li><li>售后服务：9.0</li>'
        reader = mock.Mock()
        reader.read.side_effect = [
            {'text': json.dumps({'result': True, 'html': header})},
            {'text': '公司名称：甲公司\n法定代表人：张三\n成立日期：2020-01-01\n注册资本：10万元\n电话：010-12345678\n手机：13800138000\n邮箱：example@example.com\n注册地址：测试地址'},
        ]
        reasons, evidence = [], {}
        jd.enrich_shop(reader, page, row, reasons, evidence)
        self.assertEqual(row['商品评价'], '7.2')
        self.assertEqual(row['公司名'], '甲公司')
        self.assertEqual(row['公司地址'], '测试地址')
        self.assertEqual(len(evidence), 11)
        self.assertTrue(evidence['公司名'].endswith('showLicence-123.html'))

    def test_verification_timeout_preserves_current_base_data(self):
        url = 'https://mall.jd.com/index-123.html'
        html = '<title>测试店 - 京东</title><input id="shop_id" value="123"><input id="vender_id" value="456">'
        with tempfile.TemporaryDirectory() as directory, \
             mock.patch.object(jd, 'read_shop_urls', return_value=[url, 'https://mall.jd.com/index-789.html']), \
             mock.patch.object(jd, 'JdPageReader') as reader, \
             mock.patch.object(jd, 'enrich_shop', side_effect=VerificationTimeout('验证超时')):
            reader.return_value.__enter__.return_value.read.return_value = {'html': html, 'text': ''}
            rows = jd.run_shops(outdir=directory)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]['VenderId'], '456')
            report = json.loads(next(Path(directory).glob('*.json')).read_text(encoding='utf-8'))
            self.assertIn('验证超时', report[0]['原因'])


class VerificationTests(unittest.TestCase):
    def reader(self, pages, timeout=900):
        reader = JdPageReader(verification_timeout=timeout)
        reader.session = mock.Mock()
        reader.session.wait_response.return_value = {'result': {}}
        reader._snapshot = mock.Mock(side_effect=pages)
        reader._pause = mock.Mock()
        return reader

    def test_manual_verification_resumes_without_refresh(self):
        url = 'https://mall.jd.com/index-123.html'
        ready = {'url': url, 'ready': True, 'html': '<title>店铺</title>', 'text': '店铺'}
        reader = self.reader([{'blocked': True}, {'blocked': True}, ready, ready, ready])
        with mock.patch('jd_session.time.monotonic', side_effect=itertools.count()):
            self.assertEqual(reader.read(url), ready)
        nav = [call for call in reader.session.send.call_args_list if call.args[0] == 'Page.navigate']
        self.assertEqual(len(nav), 1)

    def test_login_homepage_gets_one_return_navigation(self):
        url = 'https://mall.jd.com/index-123.html'
        ready = {'url': url, 'ready': True, 'html': '<title>店铺</title>', 'text': '店铺'}
        homepage = {'url': 'https://www.jd.com/', 'ready': True}
        reader = self.reader([{'blocked': True}, homepage, homepage, ready, ready, ready])
        with mock.patch('jd_session.time.monotonic', side_effect=itertools.count()):
            reader.read(url)
        nav = [call for call in reader.session.send.call_args_list if call.args[0] == 'Page.navigate']
        self.assertEqual(len(nav), 2)

    def test_timeout_and_stop(self):
        reader = self.reader([{'blocked': True}] * 5, timeout=2)
        with mock.patch('jd_session.time.monotonic', side_effect=itertools.count()):
            with self.assertRaises(VerificationTimeout):
                reader.read('https://mall.jd.com/index-123.html')
        stop = threading.Event()
        stop.set()
        reader = JdPageReader(stop_event=stop)
        with mock.patch.object(reader, '_connect') as connect:
            with self.assertRaises(CollectionStopped):
                reader.read('https://mall.jd.com/index-123.html')
            connect.assert_not_called()


if __name__ == '__main__':
    unittest.main()
