"""京东店铺采集与样表导出；仅填入已核实来源的字段，缺失另附报告。"""
import argparse
from datetime import datetime
from html.parser import HTMLParser
import json
from pathlib import Path
import re
import sys
import time
from urllib.parse import urlsplit

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from websocket import WebSocketException

from subject_export import RESULT_DIR
from jd_session import JdPageReader, CollectionStopped, VerificationTimeout
from jd_fields import labeled_fields, header_html, SCORE_FIELDS, COMPANY_FIELDS

FIELDS = ('店铺名', '店铺链接', 'VenderId', '店铺ID', '公司名', '法人',
          '公司注册时间', '注册资本', '电话', '手机', '邮箱', '公司地址',
          '商品评价', '物流履约', '售后服务')


def default_input_path():
    base = Path(sys.executable).parent if getattr(sys, 'frozen', False) else Path(__file__).parent
    return base / '京东店铺清单.txt'


def normalize_shop_url(value):
    parts = urlsplit(str(value).strip())
    match = re.fullmatch(r'/index-([0-9]+)\.html', parts.path)
    if (parts.scheme not in ('http', 'https') or parts.hostname != 'mall.jd.com'
            or parts.username or parts.password or parts.port not in (None, 80, 443) or not match):
        raise ValueError('请填写 https://mall.jd.com/index-数字.html 格式的京东店铺链接')
    return f'https://mall.jd.com/index-{match[1]}.html'


def read_shop_urls(path):
    """XLSX只读取店铺链接列，绝不把旧表的数据当成本次采集结果。"""
    path = Path(path)
    if path.suffix.lower() == '.xlsx':
        book = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = book.worksheets[0]
            iterator = sheet.iter_rows(values_only=True)
            headers = tuple(next(iterator, ()))
            if headers.count('店铺链接') != 1:
                raise ValueError('Excel第一行必须有且仅有一个“店铺链接”列')
            index = headers.index('店铺链接')
            values = [row[index] for row in iterator if len(row) > index and row[index]]
        finally:
            book.close()
    else:
        values = [line.strip() for line in path.read_text(encoding='utf-8-sig').splitlines()
                  if line.strip() and not line.lstrip().startswith('#')]
    urls = list(dict.fromkeys(normalize_shop_url(value) for value in values))
    if not urls:
        raise ValueError('店铺清单为空，请填写京东店铺链接')
    return urls


class ShopPageParser(HTMLParser):
    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.inputs = {}
        self.in_title = False
        self.title = []

    def handle_starttag(self, tag, attrs):
        attrs = dict(attrs)
        if tag == 'input' and attrs.get('id') in ('shop_id', 'vender_id', 'pageInstance_appId'):
            self.inputs[attrs['id']] = attrs.get('value', '').strip()
        if tag == 'title':
            self.in_title = True

    def handle_endtag(self, tag):
        if tag == 'title':
            self.in_title = False

    def handle_data(self, text):
        if self.in_title:
            self.title.append(text)


def parse_shop_page(document, url):
    url = normalize_shop_url(url)
    parser = ShopPageParser()
    parser.feed(document)
    title = ' '.join(''.join(parser.title).split())
    if '验证' in title or '欢迎登录' in title:
        raise ValueError('京东要求登录或验证，未取得店铺资料')
    shop_id = parser.inputs.get('shop_id', '')
    vender_id = parser.inputs.get('vender_id', '')
    expected = re.search(r'index-([0-9]+)', url)[1]
    if shop_id != expected or not vender_id.isascii() or not vender_id.isdigit():
        raise ValueError('页面未提供与请求店铺一致的店铺ID和VenderId')
    if not title.endswith(' - 京东'):
        raise ValueError('页面未提供可识别的店铺名称')
    row = dict.fromkeys(FIELDS, '')
    row.update({'店铺名': title.removesuffix(' - 京东').strip(), '店铺链接': url,
                '店铺ID': shop_id, 'VenderId': vender_id})
    if not row['店铺名']:
        raise ValueError('店铺名称为空')
    return row


def export_rows(rows, reports, outdir=RESULT_DIR):
    directory = Path(outdir)
    directory.mkdir(parents=True, exist_ok=True)
    stem = 'jd_shops_' + datetime.now().strftime('%Y%m%d_%H%M%S_%f')
    path = directory / (stem + '.xlsx')
    report_path = directory / (stem + '_采集说明.json')
    book = Workbook()
    sheet = book.active
    sheet.title = '商家信息'
    sheet.append(FIELDS)
    for row in rows:
        sheet.append([str(row.get(field) or '') for field in FIELDS])
    # ID、电话和外部文本均作为文本保存，避免科学计数及公式执行。
    for cells in sheet.iter_rows():
        for cell in cells:
            cell.data_type = 's'
            cell.number_format = '@'
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    widths = (30, 52, 18, 18, 36, 16, 20, 20, 22, 22, 30, 55, 14, 14, 14)
    for column, width in zip(sheet.columns, widths):
        sheet.column_dimensions[column[0].column_letter].width = width
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = sheet.dimensions
    try:
        with path.open('xb') as output:
            book.save(output)
    finally:
        book.close()
    with report_path.open('x', encoding='utf-8') as output:
        json.dump(reports, output, ensure_ascii=False, indent=2)
    print(f'[export] XLSX -> {path}', flush=True)
    print(f'[export] 采集说明 -> {report_path}', flush=True)
    return path, report_path


class HeaderTextParser(HTMLParser):
    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.parts = []
        self.hidden = 0

    def handle_starttag(self, tag, attrs):
        if tag in ('script', 'style'):
            self.hidden += 1
        if tag in ('li', 'tr', 'td', 'dt', 'dd', 'div', 'p', 'span', 'br'):
            self.parts.append('\n')

    def handle_endtag(self, tag):
        if tag in ('script', 'style'):
            self.hidden = max(0, self.hidden - 1)
        self.parts.append('\n')

    def handle_data(self, data):
        if not self.hidden:
            self.parts.append(data)


def enrich_shop(reader, page, row, reasons, evidence):
    def merge(text, fields, source):
        values, conflicts = labeled_fields(text, fields)
        for field in conflicts:
            reasons.append(f'{field}在来源页面有多个不同值，未采信：{source}')
        for field, value in values.items():
            row[field] = value
            evidence[field] = source

    merge(page['text'], SCORE_FIELDS, row['店铺链接'])
    if any(not row[field] for field in SCORE_FIELDS):
        parser = ShopPageParser()
        parser.feed(page['html'])
        app_id = parser.inputs.get('pageInstance_appId', '')
        if app_id.isascii() and app_id.isdigit():
            source = f'https://mall.jd.com/view/getJshopHeader.html?appId={app_id}'
            try:
                response = reader.read(source)
                header = HeaderTextParser()
                header.feed(header_html(response['text']))
                merge(''.join(header.parts), [field for field in SCORE_FIELDS if not row[field]], source)
            except VerificationTimeout:
                raise
            except (ValueError, RuntimeError, OSError, WebSocketException) as exc:
                reasons.append('评分读取失败：' + str(exc))
        else:
            reasons.append('店铺页面未披露评分头部appId')
    source = f'https://mall.jd.com/showLicence-{row["店铺ID"]}.html'
    try:
        license_page = reader.read(source)
        merge(license_page['text'], COMPANY_FIELDS, source)
        if not row['公司名']:
            reasons.append('资质页未取得可识别的公司名称（可能为证照图片或页面结构不同）：' + source)
    except VerificationTimeout:
        raise
    except (ValueError, RuntimeError, OSError, WebSocketException) as exc:
        reasons.append('经营资质读取失败：' + str(exc))


def run_shops(input_path=None, progress=None, outdir=RESULT_DIR, stop_event=None, verification_timeout=900):
    urls = read_shop_urls(input_path or default_input_path())
    rows, reports = [], []
    stopped = False
    def notify(stage, detail):
        if callable(progress):
            progress(len(rows), len(urls), stage, detail)
    with JdPageReader(stop_event, notify, verification_timeout) as reader:
        for index, url in enumerate(urls, 1):
            if stop_event is not None and stop_event.is_set():
                stopped = True
                break
            if index > 1:
                try:
                    if stop_event is not None:
                        if stop_event.wait(1):
                            stopped = True
                            break
                    else:
                        time.sleep(1)
                except KeyboardInterrupt:
                    stopped = True
                    break
            if callable(progress):
                progress(index - 1, len(urls), '读取京东店铺', f'{index}/{len(urls)}')
            row = dict.fromkeys(FIELDS, '')
            row['店铺链接'] = url
            reasons, evidence = [], {}
            try:
                page = reader.read(url)
                row = parse_shop_page(page['html'], url)
                evidence.update({field: url for field in FIELDS[:4]})
                enrich_shop(reader, page, row, reasons, evidence)
            except (VerificationTimeout, CollectionStopped, KeyboardInterrupt) as exc:
                stopped = True
                reasons.append(str(exc) or '用户中断，保存当前已取得的数据')
            except (ValueError, RuntimeError, OSError, WebSocketException) as exc:
                reasons.append(str(exc))
            missing = [field for field in FIELDS if not row[field]]
            if missing and not reasons:
                reasons.append('已读取的页面未披露这些字段的可识别文本；不以推测值补齐')
            rows.append(row)
            reports.append({'店铺链接': url, '采集时间': datetime.now().isoformat(timespec='seconds'),
                            '状态': ('部分完成' if missing else '完成') if row['店铺名'] else '失败',
                            '缺失列': missing, '原因': reasons, '字段来源': evidence})
            print(f'[jd] {index}/{len(urls)} {row["店铺名"] or url}：缺少{len(missing)}列', flush=True)
            if stopped:
                break
    if rows:
        export_rows(rows, reports, outdir)
    if callable(progress):
        partial = len(rows) != len(urls) or any(report['缺失列'] for report in reports)
        progress(len(rows), len(urls), '部分完成' if partial else '完成', '已保存；字段来源和缺失原因见采集说明')
    if rows and not stopped and not any(row['店铺名'] for row in rows):
        raise RuntimeError('所有店铺均未取得有效资料，失败链接及原因已保存')
    return rows


def main():
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
    parser = argparse.ArgumentParser(description='京东店铺采集，用户完成验证后继续，按样表15列输出Excel')
    parser.add_argument('--input', type=Path, default=default_input_path(), help='店铺链接TXT或含店铺链接列的XLSX')
    parser.add_argument('--outdir', type=Path, default=RESULT_DIR)
    parser.add_argument('--verification-timeout', type=int, default=900, help='等待用户登录/验证秒数，默认900')
    args = parser.parse_args()
    try:
        run_shops(input_path=args.input, outdir=args.outdir, verification_timeout=args.verification_timeout)
    finally:
        import boss_cdp
        boss_cdp.close_owned_edge()


if __name__ == '__main__':
    main()
